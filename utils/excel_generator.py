import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Tuple, Dict
import os
from datetime import datetime
import logging


class ExcelReportGenerator:
    """Generate formatted Excel reports from database query results."""
    
    def __init__(self, output_directory: str = "."):
        """
        Initialize Excel report generator.
        
        Args:
            output_directory: Directory to save Excel files
        """
        self.output_directory = output_directory
        self._validate_output_directory()
    
    def _validate_output_directory(self):
        """Validate and create output directory if needed."""
        try:
            if not os.path.exists(self.output_directory):
                os.makedirs(self.output_directory)
            
            # Test write permissions
            test_file = os.path.join(self.output_directory, "test_write.tmp")
            with open(test_file, 'w') as f:
                f.write("test")
            os.remove(test_file)
            
        except PermissionError:
            raise PermissionError(f"No write permission in directory: {self.output_directory}")
        except Exception as e:
            raise Exception(f"Cannot access output directory: {e}")
    
    def _create_styled_workbook(self, data: List[Tuple], columns: List[str], 
                              sheet_name: str) -> openpyxl.Workbook:
        """
        Create a styled Excel workbook from data.
        
        Args:
            data: Query result data as list of tuples
            columns: Column names
            sheet_name: Name for the worksheet
            
        Returns:
            Styled openpyxl Workbook
        """
        # Create DataFrame and format dates
        df = pd.DataFrame(data, columns=columns)
        
        # Convert date columns from yyyy-MM-dd to dd/MM/yyyy format
        for col in columns:
            if 'fecha' in col.lower():
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%d/%m/%Y')
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Add headers manually
        for col_idx, header in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Add data manually to have more control
        for row_idx, row_data in enumerate(df.values, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Force all date fields to be text in dd/MM/yyyy format
                if 'fecha' in columns[col_idx-1].lower():
                    cell.value = str(cell_value) if cell_value and str(cell_value) != 'NaT' else ''
                    cell.data_type = 's'  # Force string type
                else:
                    cell.value = cell_value
        
        # Style the header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Apply header styling
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Auto-adjust column widths and apply borders
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                cell.border = thin_border
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width (minimum 12, maximum 50)
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format specific data types
        self._apply_data_formatting(ws, columns)
        
        return wb
    
    def _apply_data_formatting(self, worksheet, columns: List[str]):
        """
        Apply specific formatting based on column types.
        
        Args:
            worksheet: openpyxl worksheet
            columns: List of column names
        """
        for col_idx, column_name in enumerate(columns, 1):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            
            # Format date columns - ensure they stay as text in dd/MM/yyyy format
            if 'fecha' in column_name.lower():
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"{column_letter}{row}"]
                    if cell.value and str(cell.value) != 'NaT':
                        # Ensure it's stored as text and displayed as text in dd/MM/yyyy format
                        cell.value = str(cell.value)
                        cell.data_type = 's'  # String data type
                        cell.number_format = '@'  # Text display format
            
            # Format numeric columns
            elif 'valor' in column_name.lower() or 'precio' in column_name.lower():
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"{column_letter}{row}"]
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
            
            # Center alignment for specific columns
            elif column_name.lower() in ['naturaleza_cuenta', 'tipo_empresa']:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"{column_letter}{row}"]
                    cell.alignment = Alignment(horizontal="center")
    
    def generate_invoices_report(self, data: List[Tuple], columns: List[str], 
                               start_date: str, end_date: str) -> str:
        """
        Generate invoices Excel report.
        
        Args:
            data: Invoice data as list of tuples
            columns: Column names
            start_date: Start date string
            end_date: End date string
            
        Returns:
            Path to generated Excel file
        """
        try:
            # Create filename
            if start_date == end_date:
                filename = f"facturas_{start_date}.xlsx"
            else:
                filename = f"facturas_{start_date}_{end_date}.xlsx"
            
            file_path = os.path.join(self.output_directory, filename)
            
            # Check if file exists and is in use
            self._check_file_availability(file_path)
            
            # Create styled workbook (even if data is empty)
            wb = self._create_styled_workbook(data, columns, "Facturas")
            
            # Add metadata sheet
            self._add_metadata_sheet(wb, "Facturas Report", start_date, end_date, len(data))
            
            # Save file
            wb.save(file_path)
            
            logging.info(f"Invoices report generated: {file_path}")
            return file_path
            
        except Exception as e:
            logging.error(f"Error generating invoices report: {e}")
            raise
    
    def generate_partners_report(self, data: List[Tuple], columns: List[str], 
                               start_date: str, end_date: str) -> str:
        """
        Generate partners Excel report.
        
        Args:
            data: Partners data as list of tuples
            columns: Column names
            start_date: Start date string
            end_date: End date string
            
        Returns:
            Path to generated Excel file
        """
        try:
            # Create filename
            if start_date == end_date:
                filename = f"terceros_{start_date}.xlsx"
            else:
                filename = f"terceros_{start_date}_{end_date}.xlsx"
            
            file_path = os.path.join(self.output_directory, filename)
            
            # Check if file exists and is in use
            self._check_file_availability(file_path)
            
            # Create styled workbook (even if data is empty)
            wb = self._create_styled_workbook(data, columns, "Terceros")
            
            # Add metadata sheet
            self._add_metadata_sheet(wb, "Terceros Report", start_date, end_date, len(data))
            
            # Save file
            wb.save(file_path)
            
            logging.info(f"Partners report generated: {file_path}")
            return file_path
            
        except Exception as e:
            logging.error(f"Error generating partners report: {e}")
            raise

    def generate_credit_notes_report(self, data: List[Tuple], columns: List[str], 
                                    start_date: str, end_date: str) -> str:
        """
        Generate credit notes Excel report.
        
        Args:
            data: Credit notes data as list of tuples
            columns: Column names
            start_date: Start date string
            end_date: End date string
            
        Returns:
            Path to generated Excel file
        """
        try:
            # Create filename
            if start_date == end_date:
                filename = f"notas_credito_{start_date}.xlsx"
            else:
                filename = f"notas_credito_{start_date}_{end_date}.xlsx"
            
            file_path = os.path.join(self.output_directory, filename)
            
            # Check if file exists and is in use
            self._check_file_availability(file_path)
            
            # Create styled workbook (even if data is empty)
            wb = self._create_styled_workbook(data, columns, "Notas de Crédito")
            
            # Add metadata sheet
            self._add_metadata_sheet(wb, "Credit Notes Report", start_date, end_date, len(data))
            
            # Save file
            wb.save(file_path)
            
            logging.info(f"Credit notes report generated: {file_path}")
            return file_path
            
        except Exception as e:
            logging.error(f"Error generating credit notes report: {e}")
            raise
    
    def _add_metadata_sheet(self, workbook: openpyxl.Workbook, report_type: str, 
                           start_date: str, end_date: str, record_count: int):
        """
        Add metadata information sheet to workbook.
        
        Args:
            workbook: openpyxl Workbook
            report_type: Type of report
            start_date: Start date
            end_date: End date
            record_count: Number of records
        """
        ws = workbook.create_sheet("Información del Reporte", 0)
        
        # Add metadata
        metadata = [
            ("Tipo de Reporte", report_type),
            ("Fecha de Generación", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Fecha Inicio", start_date),
            ("Fecha Fin", end_date),
            ("Total de Registros", record_count),
            ("Sistema", "Extractor de Datos Odoo 17"),
            ("Versión", "1.0")
        ]
        
        for row_idx, (key, value) in enumerate(metadata, 1):
            ws[f"A{row_idx}"] = key
            ws[f"B{row_idx}"] = value
            
            # Style the key column
            ws[f"A{row_idx}"].font = Font(bold=True)
            ws[f"A{row_idx}"].fill = PatternFill(start_color="E2E2E2", end_color="E2E2E2", fill_type="solid")
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
    
    def _check_file_availability(self, file_path: str):
        """
        Check if file is available for writing.
        
        Args:
            file_path: Path to file to check
            
        Raises:
            Exception: If file is in use or cannot be written
        """
        if os.path.exists(file_path):
            try:
                # Try to open file in write mode to check if it's in use
                with open(file_path, 'r+'):
                    pass
            except PermissionError:
                raise Exception(f"File is currently in use: {os.path.basename(file_path)}")
            except Exception as e:
                raise Exception(f"Cannot access file: {e}")
    
    def generate_combined_report(self, invoices_data: Tuple[List[Tuple], List[str]], 
                               partners_data: Tuple[List[Tuple], List[str]],
                               start_date: str, end_date: str) -> str:
        """
        Generate combined report with both invoices and partners data.
        
        Args:
            invoices_data: Tuple of (data, columns) for invoices
            partners_data: Tuple of (data, columns) for partners
            start_date: Start date string
            end_date: End date string
            
        Returns:
            Path to generated Excel file
        """
        try:
            # Create filename
            if start_date == end_date:
                filename = f"reporte_completo_{start_date}.xlsx"
            else:
                filename = f"reporte_completo_{start_date}_{end_date}.xlsx"
            
            file_path = os.path.join(self.output_directory, filename)
            
            # Check if file exists and is in use
            self._check_file_availability(file_path)
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Add metadata sheet
            self._add_metadata_sheet(wb, "Combined Report", start_date, end_date, 
                                   len(invoices_data[0]) + len(partners_data[0]))
            
            # Add invoices sheet
            if invoices_data[0]:
                inv_data, inv_columns = invoices_data
                inv_wb = self._create_styled_workbook(inv_data, inv_columns, "Facturas")
                inv_sheet = inv_wb.active
                wb._add_sheet(inv_sheet)
            
            # Add partners sheet
            if partners_data[0]:
                part_data, part_columns = partners_data
                part_wb = self._create_styled_workbook(part_data, part_columns, "Terceros")
                part_sheet = part_wb.active
                wb._add_sheet(part_sheet)
            
            # Save file
            wb.save(file_path)
            
            logging.info(f"Combined report generated: {file_path}")
            return file_path
            
        except Exception as e:
            logging.error(f"Error generating combined report: {e}")
            raise